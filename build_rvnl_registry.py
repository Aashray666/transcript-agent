"""
Build RVNL risk registry sheet and add it to risk.xlsx.

Adds a new sheet called 'RailInfrastructure' to the existing risk.xlsx
with all 57 RVNL risks, properly categorised with primary/secondary/tertiary
impact columns — exactly matching the format of all other sheets.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

# ---------------------------------------------------------------------------
# RVNL Risk Data
# Format: (Risk Name, Primary Impact, Pri Strength, Secondary Impact, Sec Strength, Tertiary Impact, Ter Strength)
# ---------------------------------------------------------------------------

RVNL_RISKS = [
    # ── Profitability & Liquidity Risks ─────────────────────────────────────
    ("Cost overruns due to delays in land acquisition, statutory approvals, and utility shifting",
     "Profitability & Liquidity", "High", "Production & Operations", "High", "Compliance", "Medium"),

    ("Inflation in steel, cement, fuel, and electrical equipment increasing project costs",
     "Profitability & Liquidity", "High", "Production & Operations", "Medium", "Strategy", "Low"),

    ("Cash-flow pressure from delayed milestone certifications and payments",
     "Profitability & Liquidity", "High", "Production & Operations", "Medium", "Strategy", "Low"),

    ("High working-capital requirements due to long project execution cycles",
     "Profitability & Liquidity", "High", "Strategy", "Medium", "Growth & Competition", "Low"),

    ("Dependence on government budget allocations and funding releases",
     "Profitability & Liquidity", "High", "Strategy", "High", "Production & Operations", "Medium"),

    ("Rising interest rates affecting long-gestation infrastructure projects",
     "Profitability & Liquidity", "High", "Strategy", "Low", "Growth & Competition", "Low"),

    ("Financial exposure from contractor claims, disputes, and arbitration",
     "Profitability & Liquidity", "High", "Compliance", "Medium", "Reputation & Ethics", "Low"),

    ("Currency fluctuations impacting imported equipment and technology",
     "Profitability & Liquidity", "High", "Production & Operations", "Low", "Technology", "Low"),

    # ── Strategy Risks ───────────────────────────────────────────────────────
    ("Misalignment between project commitments and execution capacity",
     "Strategy", "High", "Production & Operations", "High", "Profitability & Liquidity", "Medium"),

    ("Dependence on government policy decisions shaping long-term project pipelines",
     "Strategy", "High", "Profitability & Liquidity", "Medium", "Growth & Competition", "Low"),

    ("Over-extension due to aggressive expansion of project portfolios",
     "Strategy", "High", "Profitability & Liquidity", "Medium", "Production & Operations", "Medium"),

    ("Limited flexibility to respond to shifts in freight and passenger demand patterns",
     "Strategy", "High", "Customers", "Medium", "Growth & Competition", "Low"),

    ("Challenges in balancing speed of execution with governance and control",
     "Strategy", "High", "Compliance", "Medium", "Reputation & Ethics", "Low"),

    ("Delays in adopting modern construction, monitoring, and delivery practices",
     "Strategy", "High", "Production & Operations", "Medium", "Technology", "Low"),

    ("Increasing pressure to align projects with sustainability and net-zero goals",
     "Strategy", "High", "Compliance", "Medium", "Reputation & Ethics", "Medium"),

    # ── Customers / Stakeholder Risks ────────────────────────────────────────
    ("Loss of stakeholder confidence due to persistent project delays",
     "Customers", "High", "Reputation & Ethics", "High", "Strategy", "Medium"),

    ("Public dissatisfaction arising from prolonged construction and local disruption",
     "Customers", "High", "Reputation & Ethics", "Medium", "Society and People", "Medium"),

    ("Increased complaints related to land acquisition, resettlement, and rehabilitation",
     "Customers", "High", "Compliance", "Medium", "Reputation & Ethics", "Medium"),

    ("Reduced trust due to lack of transparency on project progress",
     "Customers", "High", "Reputation & Ethics", "High", "Strategy", "Low"),

    ("Changing expectations from logistics players, ports, and industrial users",
     "Customers", "Medium", "Strategy", "Medium", "Growth & Competition", "Low"),

    ("Scope changes during execution driven by evolving stakeholder demands",
     "Customers", "Medium", "Production & Operations", "High", "Profitability & Liquidity", "Medium"),

    # ── Production & Operations Risks ────────────────────────────────────────
    ("Delays in land acquisition and statutory/environmental clearances",
     "Production & Operations", "High", "Profitability & Liquidity", "High", "Compliance", "High"),

    ("Dependency on multiple external stakeholders for approvals",
     "Production & Operations", "High", "Strategy", "Medium", "Compliance", "Medium"),

    ("Contractor performance issues affecting quality and schedules",
     "Production & Operations", "High", "Profitability & Liquidity", "High", "Reputation & Ethics", "Medium"),

    ("Shortage of skilled engineers, project managers, and specialised contractors",
     "Production & Operations", "High", "Society and People", "High", "Strategy", "Medium"),

    ("Supply-chain disruptions affecting critical materials and equipment",
     "Production & Operations", "High", "Profitability & Liquidity", "High", "Strategy", "Medium"),

    ("Reliance on limited suppliers for signalling, electrification, and control systems",
     "Production & Operations", "High", "Technology", "Medium", "Strategy", "Low"),

    ("Disruptions from extreme weather events (floods, heatwaves)",
     "Production & Operations", "High", "Health, Safety & Environment", "High", "Profitability & Liquidity", "Medium"),

    ("Ineffective project monitoring leading to late identification of execution risks",
     "Production & Operations", "High", "Profitability & Liquidity", "Medium", "Strategy", "Medium"),

    # ── Compliance Risks ─────────────────────────────────────────────────────
    ("Delays or non-compliance in environmental and statutory clearances",
     "Compliance", "High", "Production & Operations", "High", "Profitability & Liquidity", "Medium"),

    ("Changes in procurement and tendering norms impacting contracts",
     "Compliance", "High", "Profitability & Liquidity", "Medium", "Strategy", "Low"),

    ("Non-compliance with safety standards in construction and electrification works",
     "Compliance", "High", "Health, Safety & Environment", "High", "Reputation & Ethics", "High"),

    ("Regulatory scrutiny from audit observations on cost escalations and delays",
     "Compliance", "High", "Reputation & Ethics", "Medium", "Profitability & Liquidity", "Medium"),

    ("Complex compliance requirements across multiple states and jurisdictions",
     "Compliance", "High", "Production & Operations", "Medium", "Strategy", "Low"),

    ("Weak contract governance leading to legal and regulatory exposure",
     "Compliance", "High", "Profitability & Liquidity", "Medium", "Reputation & Ethics", "Medium"),

    # ── Reputation & Ethics Risks ─────────────────────────────────────────────
    ("Public and media criticism from missed timelines and cost overruns",
     "Reputation & Ethics", "High", "Customers", "High", "Strategy", "Medium"),

    ("Reputational damage from safety incidents or environmental violations",
     "Reputation & Ethics", "High", "Health, Safety & Environment", "High", "Compliance", "High"),

    ("Allegations related to procurement processes or contractor selection",
     "Reputation & Ethics", "High", "Compliance", "High", "Strategy", "Medium"),

    ("Erosion of public trust due to poor disclosure and transparency",
     "Reputation & Ethics", "High", "Customers", "High", "Strategy", "Low"),

    ("Negative coverage related to land acquisition protests or litigation",
     "Reputation & Ethics", "High", "Compliance", "Medium", "Society and People", "Medium"),

    # ── Growth & Competition Risks ────────────────────────────────────────────
    ("Constraints in scaling execution capacity amidst expanding project pipeline",
     "Growth & Competition", "High", "Production & Operations", "High", "Strategy", "High"),

    ("Competition for experienced contractors and specialised vendors",
     "Growth & Competition", "High", "Production & Operations", "High", "Profitability & Liquidity", "Medium"),

    ("Delays in forming PPPs or SPVs due to partner onboarding challenges",
     "Growth & Competition", "High", "Strategy", "High", "Profitability & Liquidity", "Medium"),

    ("Risk of over-commitment leading to stretched resources",
     "Growth & Competition", "High", "Production & Operations", "High", "Strategy", "Medium"),

    ("Competing investments in alternative transport infrastructure (road, ports)",
     "Growth & Competition", "Medium", "Strategy", "High", "Customers", "Low"),

    # ── Health, Safety & Environment Risks ───────────────────────────────────
    ("Worker safety incidents during construction, electrification, and track works",
     "Health, Safety & Environment", "High", "Compliance", "High", "Reputation & Ethics", "High"),

    ("Non-compliance with occupational health and safety standards by contractors",
     "Health, Safety & Environment", "High", "Compliance", "High", "Reputation & Ethics", "Medium"),

    ("Environmental impact in ecologically sensitive areas",
     "Health, Safety & Environment", "High", "Compliance", "High", "Reputation & Ethics", "High"),

    ("Climate-related disruptions affecting construction schedules and safety",
     "Health, Safety & Environment", "High", "Production & Operations", "High", "Profitability & Liquidity", "Medium"),

    ("Litigation and protests linked to environmental and social impact",
     "Health, Safety & Environment", "High", "Compliance", "High", "Reputation & Ethics", "High"),

    ("Increasing regulatory scrutiny over sustainability and ESG performance",
     "Health, Safety & Environment", "Medium", "Compliance", "High", "Reputation & Ethics", "Medium"),

    # ── Technology Risks ──────────────────────────────────────────────────────
    ("Limited real-time visibility into project execution due to legacy systems",
     "Technology", "High", "Production & Operations", "High", "Strategy", "Medium"),

    ("Data inconsistencies across engineering, finance, and project systems",
     "Technology", "High", "Production & Operations", "Medium", "Compliance", "Low"),

    ("Cybersecurity vulnerabilities from increasing digitalisation",
     "Technology", "High", "Compliance", "Medium", "Reputation & Ethics", "Medium"),

    ("Dependence on specialised vendors for critical rail technologies",
     "Technology", "High", "Production & Operations", "High", "Strategy", "Medium"),

    ("Delays in technology deployment impacting project timelines",
     "Technology", "High", "Production & Operations", "High", "Profitability & Liquidity", "Medium"),

    ("Technology obsolescence over long asset life cycles",
     "Technology", "Medium", "Strategy", "Medium", "Production & Operations", "Low"),

    ("Integration challenges between new digital tools and existing infrastructure",
     "Technology", "High", "Production & Operations", "Medium", "Strategy", "Low"),

    # ── Society & People Risks ────────────────────────────────────────────────
    ("Shortage of skilled engineers, planners, and safety professionals",
     "Society and People", "High", "Production & Operations", "High", "Strategy", "Medium"),

    ("High dependence on contract labour impacting productivity and quality",
     "Society and People", "High", "Production & Operations", "High", "Profitability & Liquidity", "Medium"),

    ("Labour unrest or strikes affecting project execution",
     "Society and People", "High", "Production & Operations", "High", "Profitability & Liquidity", "Medium"),

    ("Community resistance due to land acquisition and resettlement",
     "Society and People", "High", "Production & Operations", "High", "Compliance", "High"),

    ("Workforce fatigue from prolonged project timelines",
     "Society and People", "Medium", "Production & Operations", "Medium", "Health, Safety & Environment", "Low"),

    ("Challenges in developing and retaining specialised rail infrastructure skills",
     "Society and People", "High", "Strategy", "High", "Growth & Competition", "Medium"),
]


def build_rvnl_registry():
    """Add RVNL sheet to risk.xlsx."""

    wb = openpyxl.load_workbook("risk.xlsx")

    # Remove existing RVNL sheet if present (idempotent)
    if "RailInfrastructure" in wb.sheetnames:
        del wb["RailInfrastructure"]
        print("Removed existing RailInfrastructure sheet")

    # Create new sheet
    ws = wb.create_sheet("RailInfrastructure")

    # ── Header row (copy style from Automotive sheet) ────────────────────────
    header = [
        "Risk Name",
        "Primary Impact",
        "Primary Impact Strength",
        "Secondary Impact",
        "Secondary Impact Strength",
        "Tertiary Impact",
        "Tertiary Impact Strength",
    ]

    # Style: bold, light blue fill (matching other sheets)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ── Data rows ────────────────────────────────────────────────────────────
    # Alternate row fill for readability
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_light = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    data_alignment = Alignment(vertical="top", wrap_text=True)

    # Category section header fills
    category_fills = {
        "Profitability & Liquidity": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "Strategy":                  PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "Customers":                 PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "Production & Operations":   PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        "Compliance":                PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid"),
        "Reputation & Ethics":       PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid"),
        "Growth & Competition":      PatternFill(start_color="D0E4F5", end_color="D0E4F5", fill_type="solid"),
        "Health, Safety & Environment": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        "Technology":                PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid"),
        "Society and People":        PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
    }

    for row_idx, risk in enumerate(RVNL_RISKS, start=2):
        row_fill = category_fills.get(risk[1], fill_white)
        for col_idx, value in enumerate(risk, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = data_alignment
            cell.border = thin_border

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 70   # Risk Name
    ws.column_dimensions["B"].width = 28   # Primary Impact
    ws.column_dimensions["C"].width = 22   # Primary Strength
    ws.column_dimensions["D"].width = 28   # Secondary Impact
    ws.column_dimensions["E"].width = 22   # Secondary Strength
    ws.column_dimensions["F"].width = 28   # Tertiary Impact
    ws.column_dimensions["G"].width = 22   # Tertiary Strength

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Save ─────────────────────────────────────────────────────────────────
    wb.save("risk.xlsx")
    print(f"Saved risk.xlsx with RailInfrastructure sheet ({len(RVNL_RISKS)} risks)")

    # ── Verify ───────────────────────────────────────────────────────────────
    wb2 = openpyxl.load_workbook("risk.xlsx", read_only=True)
    print(f"Sheets now: {wb2.sheetnames}")
    ws2 = wb2["RailInfrastructure"]
    rows = list(ws2.iter_rows(values_only=True))
    print(f"RailInfrastructure: {len(rows)-1} data rows")
    print(f"Header: {rows[0]}")
    print(f"First risk: {rows[1][0]} → {rows[1][1]}")
    print(f"Last risk:  {rows[-1][0]} → {rows[-1][1]}")
    wb2.close()


if __name__ == "__main__":
    build_rvnl_registry()
