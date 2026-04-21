"""Scoring Agent — produces impact and likelihood scores grounded in tables.

Every score MUST reference a specific row/criteria from the impact or
likelihood table. If the agent cannot map to a specific criterion, it
flags for human review rather than guessing.

Impact scoring follows the 10-step chain from how_consultants_calculate_impact.md.
Likelihood scoring uses the composite from the Likelihood Intelligence Agent.
"""

from __future__ import annotations

import json
import logging

import openpyxl

from riskmapper.llm_wrapper import LLMWrapper
from riskmapper.scoring.schemas import (
    CascadeScoringImpact,
    EvidenceContext,
    KnowledgeContext,
    LikelihoodAssessment,
    LikelihoodIntelligence,
    ScoredRisk,
    ScoringMemory,
    _LLMScoredRisk,
)

logger = logging.getLogger(__name__)

# Risk rating bands: inherent_score → rating
_RATING_BANDS = {
    range(1, 5): "Low",
    range(5, 10): "Medium",
    range(10, 16): "High",
    range(16, 26): "Critical",
}


def score_risk(
    evidence: EvidenceContext,
    knowledge: KnowledgeContext,
    likelihood: LikelihoodIntelligence,
    impact_table_text: str,
    likelihood_table: dict,
    memory: ScoringMemory,
    llm: LLMWrapper,
    external_intel=None,
) -> ScoredRisk:
    """Score a single risk for impact and likelihood.

    The Scoring Agent receives all context and produces the final scores.
    Impact is scored via the impact assessment table (7 dimensions).
    Likelihood uses the composite from the Likelihood Intelligence Agent
    as a strong prior, but the Scoring Agent can adjust ±1 with justification.
    """
    prompt = _build_scoring_prompt(
        evidence, knowledge, likelihood, impact_table_text,
        likelihood_table, memory, external_intel,
    )

    result = llm.call(
        prompt=prompt,
        response_model=_LLMScoredRisk,
        temperature=0.0,
        step_name=f"scoring_agent_{evidence.risk_id}",
    )

    # Enforce likelihood within ±1 of the code-computed composite
    llm_likelihood = result.likelihood_assessment.score
    composite = likelihood.composite_rounded
    if abs(llm_likelihood - composite) > 1:
        logger.warning(
            "%s: Likelihood %d is more than ±1 from composite %d. Clamping.",
            evidence.risk_id, llm_likelihood, composite,
        )
        result.likelihood_assessment.score = max(
            composite - 1, min(composite + 1, llm_likelihood)
        )
        result.likelihood_assessment.score = max(1, min(5, result.likelihood_assessment.score))

    # Validate inherent score = impact × likelihood
    expected_inherent = (
        result.impact_assessment.score * result.likelihood_assessment.score
    )
    if result.inherent_risk_score != expected_inherent:
        logger.warning(
            "%s: inherent_risk_score mismatch — LLM said %d, expected %d "
            "(I=%d × L=%d). Correcting.",
            evidence.risk_id,
            result.inherent_risk_score,
            expected_inherent,
            result.impact_assessment.score,
            result.likelihood_assessment.score,
        )
        result.inherent_risk_score = expected_inherent

    # Validate risk rating matches the score
    expected_rating = _get_rating(result.inherent_risk_score)
    if result.risk_rating != expected_rating:
        logger.warning(
            "%s: risk_rating mismatch — LLM said '%s', expected '%s' for "
            "score %d. Correcting.",
            evidence.risk_id,
            result.risk_rating,
            expected_rating,
            result.inherent_risk_score,
        )
        result.risk_rating = expected_rating

    # Validate impact score against evidence quantity
    _validate_impact_quantity(evidence.risk_id, result)

    # Build external intelligence audit trail
    ext_intel_used = None
    if external_intel:
        from riskmapper.scoring.schemas import ExternalIntelligenceUsed
        ext_intel_used = ExternalIntelligenceUsed(
            search_queries=external_intel.search_queries_used,
            recent_incidents=external_intel.recent_incidents,
            regulatory_developments=external_intel.regulatory_developments,
            market_trends=external_intel.market_trends,
            external_likelihood_signal=external_intel.external_likelihood_signal,
            sources=external_intel.sources,
            data_freshness=external_intel.data_freshness,
        )

    # Force cascade_likelihood_adjustment to null — only the cascade scorer
    # sets this in the second pass, not the LLM
    cascade = result.cascade_scoring_impact
    cascade.cascade_likelihood_adjustment = None

    return ScoredRisk(
        risk_id=evidence.risk_id,
        client_description=evidence.client_description,
        impact_assessment=result.impact_assessment,
        likelihood_assessment=result.likelihood_assessment,
        inherent_risk_score=result.inherent_risk_score,
        risk_rating=result.risk_rating,
        scoring_confidence=result.scoring_confidence,
        evidence_summary=result.evidence_summary,
        client_context_used=result.client_context_used,
        market_intelligence_used=ext_intel_used,
        consistency_notes=result.consistency_notes,
        flags_for_review=result.flags_for_review,
        cascade_scoring_impact=cascade,
    )


def _get_rating(score: int) -> str:
    """Map inherent risk score to rating band."""
    for score_range, rating in _RATING_BANDS.items():
        if score in score_range:
            return rating
    return "Critical"  # fallback for score=25


def _validate_impact_quantity(risk_id: str, result) -> None:
    """Check if the LLM's impact score is consistent with the evidence quantity.

    Extracts numbers from the evidence_quantity field and checks against
    common metric thresholds. Logs warnings for mismatches but doesn't
    auto-correct (the LLM may have valid reasons for its choice).
    """
    import re

    ia = result.impact_assessment
    qty_str = getattr(ia, "evidence_quantity", "") or ""
    score = ia.score
    metric_lower = ia.metric.lower()

    # Try to extract a number from the evidence quantity
    numbers = re.findall(r'[\d.]+', qty_str)
    if not numbers:
        return

    qty = float(numbers[0])

    # Check common metric thresholds
    if "days" in metric_lower or "day" in metric_lower:
        # Supply chain / logistics disruption measured in days
        expected = _lookup_days_score(qty)
        if expected and expected != score:
            logger.warning(
                "%s: Impact quantity '%s' = %.0f days → expected score %d "
                "but LLM scored %d. Correcting.",
                risk_id, qty_str, qty, expected, score,
            )
            ia.score = expected
            ia.level = _score_to_level(expected)
            # Recalculate inherent
            result.inherent_risk_score = expected * result.likelihood_assessment.score
            result.risk_rating = _get_rating(result.inherent_risk_score)

    elif "% of" in metric_lower or "% revenue" in metric_lower or "% decline" in metric_lower:
        # Percentage-based metrics
        expected = _lookup_pct_score(qty, metric_lower)
        if expected and expected != score:
            logger.warning(
                "%s: Impact quantity '%s' = %.1f%% → expected score %d "
                "but LLM scored %d. Correcting.",
                risk_id, qty_str, qty, expected, score,
            )
            ia.score = expected
            ia.level = _score_to_level(expected)
            result.inherent_risk_score = expected * result.likelihood_assessment.score
            result.risk_rating = _get_rating(result.inherent_risk_score)

    elif "hours" in metric_lower or "hour" in metric_lower:
        expected = _lookup_hours_score(qty)
        if expected and expected != score:
            logger.warning(
                "%s: Impact quantity '%s' = %.0f hours → expected score %d "
                "but LLM scored %d. Correcting.",
                risk_id, qty_str, qty, expected, score,
            )
            ia.score = expected
            ia.level = _score_to_level(expected)
            result.inherent_risk_score = expected * result.likelihood_assessment.score
            result.risk_rating = _get_rating(result.inherent_risk_score)


def _lookup_days_score(days: float) -> int | None:
    """Lookup score for days-based metrics (supply chain, logistics)."""
    if days < 1:
        return 1
    elif days <= 3:
        return 2
    elif days <= 7:
        return 3
    elif days <= 14:
        return 4
    else:
        return 5


def _lookup_hours_score(hours: float) -> int | None:
    """Lookup score for hours-based metrics (downtime, system failure)."""
    if hours < 1:
        return 1
    elif hours <= 4:
        return 2
    elif hours <= 12:
        return 3
    elif hours <= 48:
        return 4
    else:
        return 5


def _lookup_pct_score(pct: float, metric: str) -> int | None:
    """Lookup score for percentage-based metrics."""
    # Revenue decline thresholds
    if "revenue" in metric:
        if pct < 1:
            return 1
        elif pct <= 3:
            return 2
        elif pct <= 6:
            return 3
        elif pct <= 12:
            return 4
        else:
            return 5
    # Cost increase thresholds
    if "cost" in metric:
        if pct < 2:
            return 1
        elif pct <= 5:
            return 2
        elif pct <= 10:
            return 3
        elif pct <= 20:
            return 4
        else:
            return 5
    return None


def _score_to_level(score: int) -> str:
    """Map impact score to level name."""
    levels = {
        1: "No Impact",
        2: "Low",
        3: "Moderate",
        4: "High",
        5: "Severe/Catastrophic",
    }
    return levels.get(score, "Unknown")


def load_impact_table_text(xlsx_path: str, sector: str) -> str:
    """Load and format the impact assessment table from XLSX.

    Reads the sector sheet and formats all dimensions, sub-dimensions,
    metrics, and thresholds into a text block for the LLM prompt.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if sector not in wb.sheetnames:
        raise ValueError(
            f"Sector '{sector}' not found in impact table. "
            f"Available: {wb.sheetnames}"
        )

    ws = wb[sector]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    lines: list[str] = []
    current_dimension = ""

    for row in rows:
        if not row or all(cell is None for cell in row):
            continue

        # Check if this is a header row (contains "Impact Dimension")
        if row[0] and str(row[0]).strip() == "Impact Dimension":
            continue

        # Check if this is a score mapping section
        if row[0] and "Score Mapping" in str(row[0]):
            break

        # Data row
        dim = str(row[0]).strip() if row[0] else current_dimension
        if dim and dim != current_dimension:
            current_dimension = dim
            lines.append(f"\n=== {dim} ===")

        if len(row) >= 9:
            sub_dim = str(row[1]).strip() if row[1] else ""
            metric = str(row[2]).strip() if row[2] else ""
            unit = str(row[3]).strip() if row[3] else ""
            no_impact = str(row[4]).strip() if row[4] else ""
            low = str(row[5]).strip() if row[5] else ""
            moderate = str(row[6]).strip() if row[6] else ""
            high = str(row[7]).strip() if row[7] else ""
            severe = str(row[8]).strip() if row[8] else ""

            if metric:
                lines.append(
                    f"  [{sub_dim}] {metric} ({unit}):\n"
                    f"    1-No Impact: {no_impact}\n"
                    f"    2-Low: {low}\n"
                    f"    3-Moderate: {moderate}\n"
                    f"    4-High: {high}\n"
                    f"    5-Severe/Catastrophic: {severe}"
                )

    return "\n".join(lines)


def _build_scoring_prompt(
    evidence: EvidenceContext,
    knowledge: KnowledgeContext,
    likelihood: LikelihoodIntelligence,
    impact_table_text: str,
    likelihood_table: dict,
    memory: ScoringMemory,
    external_intel=None,
) -> str:
    """Build the comprehensive scoring prompt."""

    # Format likelihood factor scores
    factor_lines = []
    for fs in likelihood.factor_scores:
        factor_lines.append(
            f"  {fs.factor}: {fs.score}/5 — {fs.justification}"
        )
    likelihood_factors_text = "\n".join(factor_lines)

    # Format likelihood table levels
    likelihood_levels = _format_likelihood_levels(likelihood_table)

    # Format memory (compact)
    memory_lines = []
    for sr in memory.scored_risks:
        memory_lines.append(
            f"  {sr.risk_id}: {sr.client_description} | "
            f"I={sr.impact_score} L={sr.likelihood_score} "
            f"Score={sr.inherent_score} ({sr.risk_rating}) | "
            f"Dim={sr.dimension}"
        )
    memory_text = "\n".join(memory_lines) if memory_lines else "  (first risk being scored)"

    # Format knowledge context
    knowledge_text = json.dumps(knowledge.risk_relevant_context, indent=2)

    # Format external intelligence
    ext_intel_text = ""
    if external_intel:
        ext_lines = []
        ext_lines.append(f"  Signal: {external_intel.external_likelihood_signal}")
        if external_intel.recent_incidents:
            for inc in external_intel.recent_incidents[:3]:
                ext_lines.append(f"  - Incident: {inc}")
        if external_intel.regulatory_developments:
            for reg in external_intel.regulatory_developments[:3]:
                ext_lines.append(f"  - Regulatory: {reg}")
        if external_intel.market_trends:
            for trend in external_intel.market_trends[:3]:
                ext_lines.append(f"  - Trend: {trend}")
        ext_intel_text = "\n".join(ext_lines)
    else:
        ext_intel_text = "  (no external intelligence available)"

    return f"""You are a Senior Risk Scoring Agent for enterprise risk management.

TASK: Score the following risk for IMPACT and LIKELIHOOD using the assessment tables.
Every score MUST reference a specific table criterion. If you cannot map to a criterion,
flag for human review rather than guessing.

RISK BEING SCORED:
- Risk ID: {evidence.risk_id}
- Description: {evidence.client_description}
- Risk Type: {evidence.risk_type}
- Flags: {', '.join(evidence.flags) if evidence.flags else 'None'}
- Evidence Strength: {evidence.evidence_strength}

VERBATIM EVIDENCE FROM CLIENT:
{_format_quotes(evidence.verbatim_quotes)}

SURROUNDING CONTEXT:
{_format_quotes(evidence.surrounding_context)}

CASCADE EVIDENCE: {evidence.cascade_evidence or 'None'}
CROSS-RISK REFERENCES: {', '.join(evidence.cross_risk_references) if evidence.cross_risk_references else 'None'}

--- EXTERNAL MARKET INTELLIGENCE (from web search — EXTERNAL_INTELLIGENCE) ---
{ext_intel_text}

CLIENT QUESTIONNAIRE CONTEXT (grounded — do not add to this):
{knowledge_text}
Data completeness: {knowledge.completeness}

IMPORTANT: The questionnaire data above contains SPECIFIC QUANTITATIVE DATA (revenue figures,
percentages, headcounts, incident counts, etc.) that you MUST use to calibrate your impact score.
For example:
- If the questionnaire says revenue is EUR 78.4B and the risk could cause EUR 5B loss, that's ~6% = "High" on the Revenue Decline metric.
- If it says 22% of components are single-sourced, use that to assess supply chain severity.
- If it says EBITDA margin is 9.2% and declining, use that to assess financial resilience.
You MUST cite at least 2 specific data points from the questionnaire in your justification.

--- IMPACT ASSESSMENT TABLE ---
{impact_table_text}

--- LIKELIHOOD ASSESSMENT TABLE ---
{likelihood_levels}

--- LIKELIHOOD INTELLIGENCE (5-factor assessment) ---
Composite likelihood score: {likelihood.composite_rounded}/5 (raw: {likelihood.composite_score:.2f})
Confidence: {likelihood.confidence}
Factor breakdown:
{likelihood_factors_text}

--- PREVIOUSLY SCORED RISKS (for consistency) ---
{memory_text}

=== IMPACT SCORING INSTRUCTIONS ===

CRITICAL — DIMENSION SELECTION BIAS WARNING:
Do NOT default to "Financial & Growth Impact" for every risk. Most risks have financial
consequences eventually, but the PRIMARY impact dimension is where the MOST SEVERE and
IMMEDIATE consequence occurs. Examples:
- Supply chain disruption → Operating Impact (Critical Supplier Disruption, measured in Days)
- Cyber attack → Technology & Information Impact (system downtime, data breach)
- Tariffs/trade policy → Financial & Growth Impact (Cost Structure, not Revenue Decline)
- Product recall → Regulatory & Compliance Impact (or Customer & Market)
- Workforce issues → People, Health & Safety Impact
- Software capability gap → Technology & Information Impact (or Customer & Market for competitive loss)

You MUST evaluate at least 3 different dimensions before selecting one. State which dimensions
you considered and WHY you rejected the alternatives.

Follow this 10-step chain:
1. Read the risk description + evidence + client context.
2. Evaluate ALL 7 impact dimensions. For each, ask: "If this risk fully materializes, what is the MOST IMMEDIATE and SEVERE consequence in this dimension?"
3. EXPLICITLY list your top 3 candidate dimensions with a one-line rationale for each.
4. Select the PRIMARY dimension — the one where the MOST SEVERE and DIRECT consequence occurs. Financial consequences are often SECONDARY (downstream of the primary operational/regulatory/technology impact).
5. Justify why this dimension is primary and why you rejected the alternatives.
6. Select the most relevant sub-dimension and metric within that dimension.
7. EXTRACT THE EVIDENCE QUANTITY: Find the specific number from the evidence.
   Examples: "6 weeks" = 42 days, "EUR 400M" = 0.5% of EUR 78.4B revenue, "180,000 vehicles recalled"
   Put this in the evidence_quantity field. State where it came from in quantity_source.
8. MATCH THE QUANTITY TO THE TABLE: Look at the thresholds for your chosen metric and find
   which band the quantity falls into. This determines the score — NOT your judgment.
   Example: metric="Critical Supplier Disruption (Days)", quantity="42 days", table says ">14 = Severe" → score MUST be 5.
   Example: metric="Revenue Decline (%)", quantity="0.5%", table says "<1% = No Impact" → score MUST be 1.
9. CITE the specific table criteria that justify the score.
10. Check consistency with related risks in memory.

CRITICAL — QUANTITY-FIRST SCORING:
The score is determined by WHERE the evidence quantity falls in the table thresholds.
Do NOT pick a score first and then find a justification. Extract the quantity FIRST,
then look up the score from the table. If the evidence says "6 weeks disruption" and
the table says ">14 days = Severe/Catastrophic (5)", the score MUST be 5, not 4.

DIMENSION SELECTION RULES:
- If the risk describes a DISRUPTION (supply chain, production, logistics) → primary is Operating Impact
- If the risk describes a TECHNOLOGY FAILURE or capability gap → primary is Technology & Information Impact
- If the risk describes a REGULATORY/COMPLIANCE issue → primary is Regulatory & Compliance Impact
- If the risk describes WORKFORCE/LABOUR issues → primary is People, Health & Safety Impact
- If the risk describes COMPETITIVE/MARKET POSITION loss → primary is Customer & Market Impact
- If the risk describes a COST INCREASE or REVENUE DECLINE directly → primary is Financial & Growth Impact
- If the risk describes REPUTATION/BRAND damage → primary is Reputation & Ethics Impact

The financial consequence of an operational disruption is SECONDARY. Score the operational
disruption itself, not its financial downstream effect.

=== LIKELIHOOD SCORING INSTRUCTIONS ===
The Likelihood Intelligence Agent has computed a 5-factor composite score of {likelihood.composite_rounded}/5.
Your likelihood score MUST be {likelihood.composite_rounded}. You may ONLY adjust by ±1 if you
can cite a SPECIFIC piece of evidence that the 5-factor model missed. If you adjust, you MUST
state: "Adjusted from [composite] to [new] because [specific evidence]."
If you cannot cite a specific missed evidence point, use {likelihood.composite_rounded} exactly.

Evidence basis:
- CLIENT_STATED: Score is primarily based on what the client said in the interview.
- EXTERNAL_INTELLIGENCE: Score relies on external/sector knowledge.
- BOTH: Combination of client evidence and external intelligence.
- INSUFFICIENT: Not enough evidence — flag for review.

=== INHERENT RISK SCORE ===
Inherent Risk Score = Impact × Likelihood
Rating bands: 1-4=Low, 5-9=Medium, 10-15=High, 16-25=Critical

=== CRITICAL: ANTI-ANCHORING RULE ===
DO NOT give every risk the same score. The risks in this universe have DIFFERENT severity levels.
- Some risks have PROVEN historical losses (EUR 400M+ semiconductor shortage) → these deserve Impact 5
- Some risks are theoretical/emerging with no quantified impact → these should be Impact 2-3
- Some risks have strong tested controls (product safety) → Likelihood should be LOWER (2-3)
- Some risks have NO controls and are UNDERPREPARED → Likelihood should be HIGHER (4-5)
- Some risks are slow-build with no acute trigger → Likelihood 2-3
- Some risks are overnight/event-driven → Likelihood 4-5

You MUST differentiate. If your score matches the majority of previously scored risks,
you MUST explicitly justify why this risk deserves the SAME score rather than a different one.

SCORING CALIBRATION ANCHORS (use these as reference points):
- Impact 5: Proven loss >12% revenue, or >14 days supplier disruption, or plant shutdown
- Impact 4: Significant quantified impact (6-12% revenue, 7-14 days disruption)
- Impact 3: Moderate impact, partially quantified (3-6% revenue, 3-7 days disruption)
- Impact 2: Minor impact, no quantification, manageable (1-3% revenue, 1-3 days)
- Impact 1: Negligible, no evidence of material impact

- Likelihood 5: Currently materializing or imminent, no controls, UNDERPREPARED flag
- Likelihood 4: Occurred recently, active external drivers, weak controls
- Likelihood 3: Possible, some history at peers, partial controls exist
- Likelihood 2: Unlikely, no history at client, reasonable controls
- Likelihood 1: Rare, no history anywhere, strong tested controls

=== CONFIDENCE ===
- HIGH: Strong evidence, clear table mapping, consistent with related risks.
- MEDIUM: Adequate evidence, reasonable table mapping.
- LOW: Weak evidence, ambiguous table mapping, or inconsistency with related risks.
  If evidence_strength is WEAK, you MUST set confidence to LOW and add a flag_for_review.

=== CONSISTENCY CHECK ===
Compare your scores with previously scored risks. Flag if:
- A related risk has a very different score without clear justification.
- Your impact dimension differs from a closely related risk.

Score this risk INDEPENDENTLY based on its own evidence. Do not anchor to previous scores.
The memory is for consistency checking, not for copying scores."""


def _format_quotes(quotes: list[str]) -> str:
    if not quotes:
        return "  (none)"
    return "\n".join(f'  - "{q}"' for q in quotes)


def _format_likelihood_levels(table: dict) -> str:
    """Format likelihood levels for the scoring prompt."""
    lines: list[str] = []

    if "scale" in table:
        for score, data in sorted(table["scale"].items()):
            lines.append(
                f"  {score} ({data['level']}): {data['definition']}\n"
                f"    Frequency: {data['frequency_indicator']}"
            )
    elif "likelihood_levels" in table:
        for level in table["likelihood_levels"]:
            lines.append(
                f"  {level['score']} ({level['level']}): {level['description']}\n"
                f"    Frequency: {level['frequency_indicator']}"
            )

    return "\n".join(lines)
