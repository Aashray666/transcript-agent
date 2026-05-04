"""Risk registry loader — reads risk.xlsx and stores entries in ChromaDB.

Reads the specified sector sheet (13 sectors available) and embeds each
risk entry using all three impact dimensions for richer semantic matching.

Embedding text format:
  "{Primary Impact} | {Secondary Impact} | {Tertiary Impact} — {Risk Name}"

This gives ChromaDB much better signal than just Primary Impact + Risk Name.

Available sectors in risk.xlsx:
  Automotive, Education, Financial, FMCG, Healthcare, Insurance,
  Manufacturing, Mining, Oil&Gas, Retail, Telecommunication, Travel,
  RailInfrastructure
"""

from __future__ import annotations

import logging
import os

import openpyxl

logger = logging.getLogger(__name__)

# Canonical sector names exactly as they appear as sheet names in risk.xlsx
AVAILABLE_SECTORS = [
    "Automotive",
    "Education",
    "Financial",
    "FMCG",
    "Healthcare",
    "Insurance",
    "Manufacturing",
    "Mining",
    "Oil&Gas",
    "Retail",
    "Telecommunication",
    "Travel",
    "RailInfrastructure",
]

# Fuzzy aliases — maps common user inputs to canonical sheet names
_SECTOR_ALIASES: dict[str, str] = {
    # Automotive
    "auto": "Automotive",
    "automobile": "Automotive",
    "automotive": "Automotive",
    "vehicle": "Automotive",
    "ev": "Automotive",
    # Financial
    "financial": "Financial",
    "finance": "Financial",
    "banking": "Financial",
    "bank": "Financial",
    "fintech": "Financial",
    # FMCG
    "fmcg": "FMCG",
    "consumer goods": "FMCG",
    "consumer": "FMCG",
    "fmcg/retail": "FMCG",
    # Healthcare
    "healthcare": "Healthcare",
    "health": "Healthcare",
    "pharma": "Healthcare",
    "pharmaceutical": "Healthcare",
    "hospital": "Healthcare",
    # Insurance
    "insurance": "Insurance",
    "insurer": "Insurance",
    # Manufacturing
    "manufacturing": "Manufacturing",
    "industrial": "Manufacturing",
    "factory": "Manufacturing",
    # Mining
    "mining": "Mining",
    "minerals": "Mining",
    "metals": "Mining",
    # Oil & Gas
    "oil": "Oil&Gas",
    "gas": "Oil&Gas",
    "oil&gas": "Oil&Gas",
    "oil and gas": "Oil&Gas",
    "energy": "Oil&Gas",
    "petroleum": "Oil&Gas",
    # Retail
    "retail": "Retail",
    "ecommerce": "Retail",
    "e-commerce": "Retail",
    # Telecom
    "telecom": "Telecommunication",
    "telecommunication": "Telecommunication",
    "telecommunications": "Telecommunication",
    "telco": "Telecommunication",
    "mobile": "Telecommunication",
    # Travel
    "travel": "Travel",
    "tourism": "Travel",
    "hospitality": "Travel",
    "airline": "Travel",
    "aviation": "Travel",
    # Rail / Infrastructure / RVNL
    "rail": "RailInfrastructure",
    "railinfrastructure": "RailInfrastructure",
    "infrastructure": "RailInfrastructure",
    "railway": "RailInfrastructure",
    "metro": "RailInfrastructure",
    "rvnl": "RVNL",
    "rail vikas nigam": "RVNL",
    "rail vikas": "RVNL",
    "railvikas": "RVNL",
    # Education
    "education": "Education",
    "university": "Education",
    "school": "Education",
    "edtech": "Education",
}


def resolve_sector(sector_input: str, xlsx_path: str = "risk.xlsx") -> str:
    """Resolve a user-provided sector string to the canonical sheet name.

    Tries exact match first, then alias lookup, then fuzzy substring match.

    Args:
        sector_input: User-provided sector string (case-insensitive).
        xlsx_path: Path to risk.xlsx (used to read actual sheet names).

    Returns:
        Canonical sheet name (e.g. "Automotive").

    Raises:
        ValueError: If no match found, with a list of available sectors.
    """
    # Get actual sheet names from the file
    available = _get_sheet_names(xlsx_path)

    # 1. Exact match (case-insensitive)
    for sheet in available:
        if sheet.lower() == sector_input.lower():
            return sheet

    # 2. Alias lookup
    alias_key = sector_input.lower().strip()
    if alias_key in _SECTOR_ALIASES:
        canonical = _SECTOR_ALIASES[alias_key]
        if canonical in available:
            return canonical

    # 3. Fuzzy substring match
    sector_lower = sector_input.lower()
    for sheet in available:
        if sector_lower in sheet.lower() or sheet.lower() in sector_lower:
            logger.info(
                "Sector '%s' fuzzy-matched to sheet '%s'", sector_input, sheet
            )
            return sheet

    raise ValueError(
        f"Sector '{sector_input}' not found in {xlsx_path}.\n"
        f"Available sectors: {', '.join(available)}\n"
        f"Common aliases: automotive, banking, healthcare, manufacturing, "
        f"oil&gas, telecom, retail, fmcg, mining, rail, travel, insurance, education"
    )


def list_available_sectors(xlsx_path: str = "risk.xlsx") -> list[str]:
    """Return all sector sheet names available in the registry."""
    return _get_sheet_names(xlsx_path)


def load_registry(
    xlsx_path: str,
    sector: str,
    chroma_client,
    embedding_fn=None,
) -> int:
    """Load risk entries from the sector sheet into ChromaDB.

    Uses all three impact dimensions (Primary, Secondary, Tertiary) to
    build richer embedding text for better semantic matching.

    Args:
        xlsx_path: Path to the risk registry XLSX workbook (risk.xlsx).
        sector: Sheet name — use resolve_sector() if unsure of exact name.
        chroma_client: A chromadb.ClientAPI instance.
        embedding_fn: Optional ChromaDB embedding function.

    Returns:
        Number of entries loaded.

    Raises:
        FileNotFoundError: If xlsx_path does not exist.
        ValueError: If the sector sheet is not found or is empty.
    """
    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(
            f"Risk registry file not found: {xlsx_path}\n"
            f"Make sure risk.xlsx is in the project folder."
        )

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    if sector not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sector sheet '{sector}' not found in {xlsx_path}.\n"
            f"Available sheets: {wb.sheetnames}\n"
            f"Tip: use resolve_sector() to auto-match your sector name."
        )

    ws = wb[sector]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header row
    wb.close()

    # Filter out completely empty rows
    rows = [r for r in rows if r and r[0] is not None]

    if len(rows) == 0:
        raise ValueError(
            f"Sector sheet '{sector}' in {xlsx_path} contains zero risk entries."
        )

    # Build sector prefix for IDs (first 3 chars uppercase)
    sector_prefix = sector[:3].upper()

    # Idempotency: delete + recreate collection each run
    try:
        chroma_client.delete_collection("risk_registry")
    except Exception:
        pass

    collection_kwargs = {"name": "risk_registry"}
    if embedding_fn is not None:
        collection_kwargs["embedding_function"] = embedding_fn
    collection = chroma_client.create_collection(**collection_kwargs)

    # Update available sectors list to include RVNL
    global AVAILABLE_SECTORS
    if "RVNL" not in AVAILABLE_SECTORS:
        AVAILABLE_SECTORS = _get_sheet_names(xlsx_path)

    ids = []
    documents = []
    metadatas = []

    # Forward-fill Primary Impact — RVNL sheet only fills the first row
    # of each category group; the rest are None (merged-cell style)
    last_primary_impact = ""

    for idx, row in enumerate(rows, start=1):
        # Only use col 0 (Risk Name) and col 1 (Primary Impact)
        risk_name      = _cell(row, 0)
        primary_impact = _cell(row, 1)

        if not risk_name:
            continue

        # Forward-fill: if col 1 is blank, carry forward the last seen value
        if primary_impact:
            last_primary_impact = primary_impact
        else:
            primary_impact = last_primary_impact

        registry_risk_id = f"REG_{sector_prefix}_{idx:03d}"

        # Embedding text: "Primary Impact - Risk Name"
        doc_text = f"{primary_impact} - {risk_name}" if primary_impact else risk_name

        ids.append(registry_risk_id)
        documents.append(doc_text)
        metadatas.append({
            "registry_risk_id": registry_risk_id,
            "risk_name": risk_name,
            "primary_impact": primary_impact,
            "sector": sector,
        })

    if not ids:
        raise ValueError(
            f"No valid risk entries found in sheet '{sector}' of {xlsx_path}."
        )

    collection.add(ids=ids, documents=documents, metadatas=metadatas)

    logger.info(
        "Registry loaded | sector=%s | entries=%d | collection=risk_registry",
        sector, len(ids),
    )

    return len(ids)


# ── helpers ───────────────────────────────────────────────────────────────────

def _cell(row: tuple, idx: int) -> str:
    """Safely extract and clean a cell value from a row tuple."""
    if idx >= len(row) or row[idx] is None:
        return ""
    val = str(row[idx]).strip()
    return val if val.lower() not in ("none", "n/a", "-", "") else ""


def _get_sheet_names(xlsx_path: str) -> list[str]:
    """Get sheet names from the workbook without loading all data."""
    if not os.path.isfile(xlsx_path):
        return AVAILABLE_SECTORS  # fallback to hardcoded list
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return AVAILABLE_SECTORS
