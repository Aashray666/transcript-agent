"""
Phase 1 — Document Risk Extraction Pipeline
============================================
Upload any document (PDF, DOCX, XLSX, TXT) and this script:
  1. Loads the document
  2. Chunks it (section-aware, overlapping)
  3. Extracts every risk mention per chunk via LLM
  4. Deduplicates across chunks
  5. Maps each risk to the RVNL / sector registry (risk.xlsx)
  6. Writes Excel output with 4 sheets

SETUP:
  Open .env and set:  NVIDIA_API_KEY=nvapi-YOUR_KEY_HERE

USAGE (single document):
  python3 run_phase1_from_document.py --file report.pdf --sector rvnl

USAGE (multiple documents, same output file):
  python3 run_phase1_from_document.py --file doc1.pdf --sector rvnl
  python3 run_phase1_from_document.py --file doc2.pdf --sector rvnl --append

DOCUMENT TYPES:
  annual_report  audit_report  con_call_transcript
  board_presentation  management_letter  regulatory_filing  other
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import time

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("phase1")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Phase 1: Extract and map risks from any document",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--file", "-f", required=True,
                   help="Path to document (PDF, DOCX, XLSX, TXT, MD)")
    p.add_argument("--type", "-t", default="other",
                   choices=["audit_report","con_call_transcript","annual_report",
                            "board_presentation","management_letter",
                            "regulatory_filing","other"],
                   help="Document type. Default: other")
    p.add_argument("--sector", "-s", default=None,
                   help="Sector: rvnl, automotive, banking, healthcare, etc.")
    p.add_argument("--registry", "-r", default="risk.xlsx",
                   help="Risk registry XLSX. Default: risk.xlsx")
    p.add_argument("--output", "-o", default="output_phase1",
                   help="Output directory. Default: output_phase1")
    p.add_argument("--skip-registry", action="store_true",
                   help="Skip registry mapping (just extract risks)")
    p.add_argument("--append", action="store_true",
                   help="Append to existing Excel output instead of overwriting")
    p.add_argument("--max-chunk-chars", type=int, default=3000)
    p.add_argument("--overlap-chars", type=int, default=200)
    p.add_argument("--max-chunks", type=int, default=None,
                   help="Only process first N chunks (useful for large docs). Default: all")
    p.add_argument("--start-chunk", type=int, default=0,
                   help="Skip first N chunks (useful to skip empty pages). Default: 0")
    return p.parse_args()


# ---------------------------------------------------------------------------
# API key guard
# ---------------------------------------------------------------------------

def _check_api_key() -> None:
    from dotenv import load_dotenv
    load_dotenv(override=True)
    key = os.environ.get("NVIDIA_API_KEY", "")
    if not key or key == "nvapi-PASTE_YOUR_KEY_HERE":
        print()
        print("=" * 60)
        print("  API KEY NOT SET")
        print("=" * 60)
        print("  Open the file  .env  and replace:")
        print("      NVIDIA_API_KEY=nvapi-PASTE_YOUR_KEY_HERE")
        print("  with your actual key from https://build.nvidia.com")
        print()
        sys.exit(1)
    logger.info("API key loaded (ends ...%s)", key[-6:])


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    args = _parse_args()

    # Validate file
    if not os.path.isfile(args.file):
        logger.error("File not found: %s", args.file)
        sys.exit(1)

    ext = os.path.splitext(args.file)[1].lower()
    if ext not in {".pdf", ".docx", ".xlsx", ".txt", ".md"}:
        logger.error("Unsupported format '%s'. Use: pdf, docx, xlsx, txt, md", ext)
        sys.exit(1)

    _check_api_key()

    try:
        import chromadb
        from riskmapper.document_extractor.extractor import extract_risks_from_document
        from riskmapper.llm_wrapper import LLMWrapper
        from riskmapper.risk_registry_loader import resolve_sector
    except ImportError as exc:
        logger.error("Import error: %s — run: pip3 install -r requirements.txt", exc)
        sys.exit(1)

    os.makedirs(args.output, exist_ok=True)
    doc_name = os.path.basename(args.file)
    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in doc_name)[:50]

    # Resolve sector
    registry_exists = os.path.isfile(args.registry)
    if args.sector and registry_exists:
        try:
            sector = resolve_sector(args.sector, args.registry)
        except ValueError as exc:
            print(f"\nERROR: {exc}\n")
            sys.exit(1)
    elif args.sector:
        sector = args.sector
    else:
        sector = _pick_sector_interactively(args.registry)

    print()
    print("=" * 60)
    print("  PHASE 1 — DOCUMENT RISK EXTRACTION")
    print("=" * 60)
    print(f"  Document : {doc_name}")
    print(f"  Type     : {args.type}")
    print(f"  Sector   : {sector}")
    print(f"  Registry : {args.registry}")
    print(f"  Output   : {args.output}/")
    print("=" * 60)
    print()

    t0 = time.time()
    llm = LLMWrapper()

    # ------------------------------------------------------------------
    # STEP 1-4: Load → Chunk → Extract → Deduplicate
    # ------------------------------------------------------------------
    logger.info("STEP 1-4: Load, chunk, extract, deduplicate...")
    extraction = extract_risks_from_document(
        file_path=args.file,
        document_type=args.type,
        sector=sector,
        llm=llm,
        output_dir=None,
        max_chunk_chars=args.max_chunk_chars,
        overlap_chars=args.overlap_chars,
        max_chunks=args.max_chunks,
        start_chunk=args.start_chunk,
    )

    deduped = extraction.risks
    s = extraction.summary

    if not deduped:
        print("\n  No risks extracted. Check document content and try again.")
        sys.exit(0)

    logger.info("Extracted %d unique risks from %d raw mentions",
                s.deduplicated_risks, s.raw_mentions)

    # ------------------------------------------------------------------
    # STEP 5: Registry Mapping
    # ------------------------------------------------------------------
    mapped_risks = None
    if not args.skip_registry and registry_exists:
        logger.info("STEP 5: Mapping %d risks to registry...", len(deduped))
        try:
            mapped_risks = _run_registry_mapping(deduped, sector, args.registry, llm)
            n_mapped   = sum(1 for r in mapped_risks if not r.unmapped)
            n_unmapped = sum(1 for r in mapped_risks if r.unmapped)
            logger.info("Registry mapping done | mapped=%d | unmapped=%d",
                        n_mapped, n_unmapped)
        except Exception as exc:
            logger.warning("Registry mapping failed: %s — skipping", exc)
    elif not args.skip_registry:
        logger.warning("Registry file not found: %s", args.registry)

    # ------------------------------------------------------------------
    # STEP 6: Write outputs (JSON + Excel)
    # ------------------------------------------------------------------
    excel_path = os.path.join(args.output, "risk_extraction_output.xlsx")
    _write_excel(
        deduped_risks=deduped,
        mapped_risks=mapped_risks,
        doc_name=doc_name,
        doc_type=args.type,
        sector=sector,
        excel_path=excel_path,
        append=args.append,
    )

    # Also write JSON for downstream pipeline compatibility
    json_path = os.path.join(args.output, f"{safe_name}_risks.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump([r.model_dump(mode="json") for r in deduped],
                  f, indent=2, ensure_ascii=False)

    if mapped_risks:
        mapped_json = os.path.join(args.output, f"{safe_name}_registry_mapped.json")
        with open(mapped_json, "w", encoding="utf-8") as f:
            json.dump([r.model_dump(mode="json") for r in mapped_risks],
                      f, indent=2, ensure_ascii=False)

    elapsed = time.time() - t0

    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------
    print()
    print("=" * 60)
    print(f"  COMPLETE — {s.deduplicated_risks} risks extracted in {elapsed:.0f}s")
    print("=" * 60)
    print(f"  Chunks processed : {s.total_chunks}")
    print(f"  Raw mentions     : {s.raw_mentions}")
    print(f"  After dedup      : {s.deduplicated_risks}")
    print()
    print(f"  Severity:  Critical={s.critical_count}  High={s.high_count}"
          f"  Medium={s.medium_count}  Low={s.low_count}")
    if s.material_weakness_count:
        print(f"  Material weaknesses : {s.material_weakness_count}")
    if s.repeat_finding_count:
        print(f"  Repeat findings     : {s.repeat_finding_count}")
    if s.quantified_count:
        print(f"  Quantified risks    : {s.quantified_count}")
    print()
    print(f"  Output: {excel_path}")
    print()

    _sev_rank = {"CRITICAL":5,"HIGH":4,"MEDIUM":3,"LOW":2,"UNSPECIFIED":1}
    top = sorted(deduped, key=lambda r: -_sev_rank.get(r.severity_signal, 1))
    print("  Top risks:")
    for r in top[:8]:
        qty   = f" [{r.financial_quantification}]" if r.financial_quantification else ""
        flags = f" | {', '.join(r.flags[:2])}" if r.flags else ""
        print(f"    {r.risk_id} [{r.severity_signal:11}] {r.client_description[:55]}{qty}{flags}")
    print()


# ---------------------------------------------------------------------------
# Registry mapping bridge
# ---------------------------------------------------------------------------

def _run_registry_mapping(deduped_risks, sector, registry_path, llm):
    import chromadb
    from riskmapper.risk_registry_loader import load_registry
    from riskmapper.registry_mapper import map_risks
    from riskmapper.schemas import DeduplicatedRisk

    chroma_client = chromadb.Client()
    load_registry(registry_path, sector, chroma_client)

    phase1_risks = []
    for i, dr in enumerate(deduped_risks):
        # Map document flags → Phase 1 flags
        flags = []
        if "MATERIAL_WEAKNESS" in dr.flags or "SIGNIFICANT_DEFICIENCY" in dr.flags:
            flags.append("UNDERPREPARED")
        if "REPEAT_FINDING" in dr.flags:
            flags.append("UNREGISTERED")
        if "CASCADE_SIGNAL" in dr.flags:
            flags.append("CASCADE_SIGNAL")

        phase1_risks.append(DeduplicatedRisk(
            risk_id=dr.risk_id.replace("DOC_RISK_", "RISK_"),
            client_description=dr.client_description,
            verbatim_evidence=dr.verbatim_evidence[:5],
            question_source=[f"Q{(i % 18) + 1}"],
            risk_type=dr.risk_type,
            flags=flags,
            cascade_context=dr.cascade_context,
            merged_from=dr.merged_from,
        ))

    return map_risks(phase1_risks, sector, chroma_client, llm)


# ---------------------------------------------------------------------------
# Excel output — 4 sheets
# ---------------------------------------------------------------------------

def _write_excel(
    deduped_risks,
    mapped_risks,
    doc_name: str,
    doc_type: str,
    sector: str,
    excel_path: str,
    append: bool,
) -> None:
    """Write (or append to) the Excel output file.

    Sheet 1 — Risk Register     : one row per risk, all key fields
    Sheet 2 — Evidence          : verbatim quotes, one row per quote
    Sheet 3 — Registry Mapping  : registry matches with confidence scores
    Sheet 4 — Review Queue      : risks flagged for human review
    """
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter

    _sev_rank = {"CRITICAL":5,"HIGH":4,"MEDIUM":3,"LOW":2,"UNSPECIFIED":1}

    # Severity → fill colour (Excel ARGB)
    _sev_fill = {
        "CRITICAL":    "FFFF0000",   # red
        "HIGH":        "FFFF6600",   # orange
        "MEDIUM":      "FFFFFF00",   # yellow
        "LOW":         "FF92D050",   # green
        "UNSPECIFIED": "FFD9D9D9",   # grey
    }

    # ── load or create workbook ───────────────────────────────────────────────
    if append and os.path.isfile(excel_path):
        wb = openpyxl.load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # ── helper: get or create sheet ───────────────────────────────────────────
    def _sheet(name: str, headers: list[str]) -> openpyxl.worksheet.worksheet.Worksheet:
        if name in wb.sheetnames:
            ws = wb[name]
        else:
            ws = wb.create_sheet(name)
            # Write header row
            ws.append(headers)
            # Style header
            hdr_fill = PatternFill("solid", fgColor="FF1F3864")
            hdr_font = Font(bold=True, color="FFFFFFFF", size=10)
            hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = hdr_align
            ws.row_dimensions[1].height = 30
        return ws

    # ── helper: style a data row ──────────────────────────────────────────────
    thin = Side(style="thin", color="FFD9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _style_row(ws, row_num: int, sev: str = "UNSPECIFIED") -> None:
        fill_color = _sev_fill.get(sev, "FFFFFFFF")
        fill = PatternFill("solid", fgColor=fill_color)
        for cell in ws[row_num]:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column == 1:   # severity column gets colour
                cell.fill = fill

    # ── build lookup: doc_risk_id → mapped_risk ───────────────────────────────
    mapped_lookup: dict[str, object] = {}
    if mapped_risks:
        for mr in mapped_risks:
            # match by description (bridge between schemas)
            for dr in deduped_risks:
                if dr.client_description == mr.client_description:
                    mapped_lookup[dr.risk_id] = mr
                    break

    sorted_risks = sorted(
        deduped_risks,
        key=lambda r: -_sev_rank.get(r.severity_signal, 1),
    )

    # =========================================================================
    # SHEET 1 — Risk Register
    # =========================================================================
    ws1 = _sheet("Risk Register", [
        "Severity", "Risk ID", "Risk Description",
        "Category", "Risk Type", "Source Document",
        "Source Sections", "Occurrence Count",
        "Financial Quantification", "Management Response",
        "Flags", "Mapped Registry Risk", "Registry Category",
        "Match Confidence", "Unmapped",
    ])

    for dr in sorted_risks:
        mr = mapped_lookup.get(dr.risk_id)

        # Best registry match
        reg_risk = reg_cat = conf = ""
        unmapped = "Yes"
        if mr:
            unmapped = "Yes" if mr.unmapped else "No"
            if mr.registry_matches:
                best = mr.registry_matches[0]
                reg_risk = best.risk_name
                reg_cat  = best.primary_impact
                conf     = best.confidence

        ws1.append([
            dr.severity_signal,
            dr.risk_id,
            dr.client_description,
            dr.risk_category,
            dr.risk_type,
            doc_name,
            " | ".join(dr.source_sections[:3]),
            dr.occurrence_count,
            dr.financial_quantification or "",
            dr.management_response or "",
            ", ".join(dr.flags),
            reg_risk,
            reg_cat,
            conf,
            unmapped,
        ])
        _style_row(ws1, ws1.max_row, dr.severity_signal)

    # Column widths for sheet 1
    for col, width in zip("ABCDEFGHIJKLMNO",
                          [12,12,55,20,14,30,30,8,25,35,30,45,22,12,10]):
        ws1.column_dimensions[col].width = width

    # =========================================================================
    # SHEET 2 — Evidence
    # =========================================================================
    ws2 = _sheet("Evidence", [
        "Risk ID", "Risk Description", "Severity",
        "Source Document", "Source Section", "Verbatim Evidence",
    ])

    for dr in sorted_risks:
        for quote in dr.verbatim_evidence:
            ws2.append([
                dr.risk_id,
                dr.client_description,
                dr.severity_signal,
                doc_name,
                " | ".join(dr.source_sections[:2]),
                quote,
            ])
            _style_row(ws2, ws2.max_row, dr.severity_signal)

    for col, width in zip("ABCDEF", [12, 45, 12, 30, 30, 80]):
        ws2.column_dimensions[col].width = width

    # =========================================================================
    # SHEET 3 — Registry Mapping
    # =========================================================================
    ws3 = _sheet("Registry Mapping", [
        "Risk ID", "Extracted Risk Description",
        "Registry Risk ID", "Registry Risk Name",
        "Primary Impact Category", "Similarity Score",
        "Match Confidence", "Unmapped", "Source Document",
    ])

    for dr in sorted_risks:
        mr = mapped_lookup.get(dr.risk_id)
        if mr and mr.registry_matches:
            for match in mr.registry_matches[:3]:
                ws3.append([
                    dr.risk_id,
                    dr.client_description,
                    match.registry_risk_id,
                    match.risk_name,
                    match.primary_impact,
                    round(match.similarity_score, 3),
                    match.confidence,
                    "Yes" if mr.unmapped else "No",
                    doc_name,
                ])
                _style_row(ws3, ws3.max_row)
        else:
            ws3.append([
                dr.risk_id,
                dr.client_description,
                "—", "—", "—", "—", "—",
                "Yes (no registry)",
                doc_name,
            ])
            _style_row(ws3, ws3.max_row)

    for col, width in zip("ABCDEFGHI", [12, 55, 14, 55, 25, 12, 14, 10, 30]):
        ws3.column_dimensions[col].width = width

    # =========================================================================
    # SHEET 4 — Review Queue
    # =========================================================================
    ws4 = _sheet("Review Queue", [
        "Risk ID", "Risk Description", "Severity",
        "Reason for Review", "Source Document",
        "Flags", "Financial Quantification",
    ])

    review_risks = [
        dr for dr in sorted_risks
        if (mapped_lookup.get(dr.risk_id) and mapped_lookup[dr.risk_id].human_review)
        or "MATERIAL_WEAKNESS" in dr.flags
        or "UNMITIGATED" in dr.flags
    ]

    for dr in review_risks:
        mr = mapped_lookup.get(dr.risk_id)
        reason = (mr.human_review_reason if mr and mr.human_review_reason
                  else "Flagged: " + ", ".join(dr.flags[:2]))
        ws4.append([
            dr.risk_id,
            dr.client_description,
            dr.severity_signal,
            reason,
            doc_name,
            ", ".join(dr.flags),
            dr.financial_quantification or "",
        ])
        _style_row(ws4, ws4.max_row, dr.severity_signal)

    for col, width in zip("ABCDEFG", [12, 55, 12, 40, 30, 30, 25]):
        ws4.column_dimensions[col].width = width

    # ── freeze top row on all sheets ─────────────────────────────────────────
    for ws in [ws1, ws2, ws3, ws4]:
        ws.freeze_panes = "A2"

    wb.save(excel_path)
    logger.info("Excel written: %s  (%d risks, %d evidence rows, %d review)",
                excel_path, len(sorted_risks),
                sum(len(dr.verbatim_evidence) for dr in deduped_risks),
                len(review_risks))


# ---------------------------------------------------------------------------
# Interactive sector picker
# ---------------------------------------------------------------------------

def _pick_sector_interactively(registry_path: str) -> str:
    from riskmapper.risk_registry_loader import list_available_sectors
    sectors = list_available_sectors(registry_path)
    print()
    print("  Available sectors in risk.xlsx:")
    for i, s in enumerate(sectors, 1):
        print(f"    {i:2d}. {s}")
    print()
    while True:
        choice = input("  Enter sector name or number: ").strip()
        if choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(sectors):
                return sectors[idx]
        else:
            from riskmapper.risk_registry_loader import resolve_sector
            try:
                return resolve_sector(choice, registry_path)
            except ValueError:
                pass
        print(f"  Invalid choice '{choice}'. Try again.")


if __name__ == "__main__":
    main()
